"""Lectura y plantilla de Excel para carga masiva de equipos (RF-008).

El parseo es POSICIONAL: se usa el orden de las columnas, no el texto del
encabezado. Por eso ENCABEZADOS (lo que ve el usuario) y COLUMNAS (las claves
internas) van alineados 1 a 1. Los campos de selección múltiple se escriben en
una sola celda separados por ';' (ej.: "Operación; Mantenimiento").
"""
from datetime import date, datetime
from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

# (clave_interna, encabezado_visible, valor_de_ejemplo)
_DEFINICION: list[tuple[str, str, str]] = [
    ("codigo_interno", "Código interno", "EQ-0001"),
    ("serial_fabricante", "Serial", "SN12345"),
    ("nombre", "Nombre", "Monitor de signos vitales"),
    ("marca", "Marca", "Mindray"),
    ("modelo", "Modelo", "uMEC12"),
    ("numero_activo", "N° de activo (placa)", "AF-00123"),
    ("estado", "Estado", "operativo"),
    ("sede", "Sede", "Sede Principal"),
    ("servicio", "Servicio", "UCI"),
    ("piso", "Piso", "3"),
    ("clase_biomedica", "Clase biomédica", "Diagnostico"),
    ("clase_uso", "Clase de uso", "Medico"),
    ("clasificacion_riesgo", "Clasificación según riesgo", "IIb"),
    ("tecnologia_predominante", "Tecnología predominante", "Electrónica"),
    ("fabricante", "Fabricante", "Mindray"),
    ("anio_fabricacion", "Año de fabricación", "2022"),
    ("pais_fabricante", "País fabricante", "China"),
    ("ciudad_fabricante", "Ciudad fabricante", "Shenzhen"),
    ("direccion_fabricante", "Dirección fabricante", ""),
    ("telefono_fabricante", "Teléfono fabricante", ""),
    ("correo_fabricante", "Correo fabricante", ""),
    ("representante", "Representante", ""),
    ("pais_representante", "País representante", "Colombia"),
    ("ciudad_representante", "Ciudad representante", "Bogotá"),
    ("direccion_representante", "Dirección representante", ""),
    ("telefono_representante", "Teléfono representante", ""),
    ("correo_representante", "Correo representante", ""),
    ("voltaje_operacion", "Voltaje de operación", "110 V"),
    ("voltaje_maximo", "Voltaje máximo", "240 V"),
    ("corriente_maxima", "Corriente máxima", ""),
    ("corriente_minima", "Corriente mínima", ""),
    ("potencia_consumida", "Potencia consumida", ""),
    ("frecuencia", "Frecuencia", "60 Hz"),
    ("presion", "Presión", ""),
    ("velocidad", "Velocidad", ""),
    ("temperatura", "Temperatura", ""),
    ("peso", "Peso", ""),
    ("capacidad", "Capacidad", ""),
    ("fuentes_alimentacion", "Fuente de alimentación (separar con ;)", "Electricidad"),
    ("manuales", "Manuales (separar con ;)", "Operación; Mantenimiento"),
    ("planos", "Planos (separar con ;)", "Ninguno"),
    ("recomendaciones_fabricante", "Recomendaciones fabricante (separar con ;)", ""),
    ("modo_adquisicion", "Modo de adquisición", "Compra Directa"),
    ("propiedad", "Propiedad", "propio"),
    ("proveedor", "Proveedor", "Proveedor Uno"),
    ("fecha_adquisicion", "Fecha adquisición", "2025-01-15"),
    ("costo_adquisicion", "Costo adquisición", "12500000"),
    ("orden_compra", "Orden de compra", "OC-77"),
    ("fecha_inicial_garantia", "Fecha inicial garantía", "2025-01-15"),
    ("fecha_final_garantia", "Fecha final garantía", "2027-01-15"),
    ("fecha_instalacion", "Fecha instalación", ""),
    ("fecha_funcionamiento", "Fecha funcionamiento", ""),
    ("registro_invima", "Registro INVIMA", "INVIMA-2020-001"),
    ("fecha_vencimiento_invima", "Fecha vencimiento INVIMA", "2028-01-15"),
    ("periodicidad_mantenimiento", "Periodicidad de mantenimiento", "Semestral"),
    ("calibracion_si", "Calibración SÍ", "Sí"),
    ("calibracion_no", "Calibración NO", ""),
    ("equipo_movil", "Equipo móvil", "Sí"),
    ("equipo_fijo", "Equipo fijo", ""),
    ("accesorios", "Accesorios", ""),
    ("descripcion_funcional", "Descripción funcional", ""),
]

COLUMNAS: list[str] = [clave for clave, _, _ in _DEFINICION]
ENCABEZADOS: list[str] = [enc for _, enc, _ in _DEFINICION]
_EJEMPLO: list[str] = [ej for _, _, ej in _DEFINICION]


def _celda_a_texto(valor: object) -> str | None:
    if valor is None:
        return None
    if isinstance(valor, datetime | date):
        return valor.date().isoformat() if isinstance(valor, datetime) else valor.isoformat()
    texto = str(valor).strip()
    return texto or None


def parse_equipos(contenido: bytes) -> list[tuple[int, dict[str, str | None]]]:
    """Devuelve una lista de (número de fila, {columna: valor}) desde la 2ª fila."""
    wb = load_workbook(BytesIO(contenido), data_only=True)
    ws = wb.active
    filas: list[tuple[int, dict[str, str | None]]] = []
    for indice, fila in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        valores = {
            col: _celda_a_texto(fila[i] if i < len(fila) else None)
            for i, col in enumerate(COLUMNAS)
        }
        if any(v is not None for v in valores.values()):
            filas.append((indice, valores))
    return filas


def generar_plantilla() -> bytes:
    """Genera un Excel con la fila de encabezados y un ejemplo."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Equipos"
    ws.append(ENCABEZADOS)
    for celda in ws[1]:
        celda.font = Font(bold=True)
    ws.append(_EJEMPLO)
    salida = BytesIO()
    wb.save(salida)
    return salida.getvalue()
